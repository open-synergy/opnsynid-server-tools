# -*- coding: utf-8 -*-
# Copyright 2020 OpenSynergy Indonesia
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
import re
import os
import time
from openerp.tools.float_utils import float_compare
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError
from io import BytesIO
import base64
import logging
from datetime import date, datetime as dt
_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.styles import colors, PatternFill, Alignment, Font
except ImportError:
    _logger.debug(
        "Cannot import 'openpyxl'. Please make sure it is installed.")


class BaseExportXlsx(models.AbstractModel):
    _name = "base.export.xlsx"

    @api.model
    def get_eval_context(self, model, record, value):
        eval_context = {
            "float_compare": float_compare,
            "time": time,
            "datetime": dt,
            "date": date,
            "value": value,
            "object": record,
            "model": self.env[model],
            "env": self.env,
            "context": self._context,
        }
        return eval_context

    @api.model
    def get_styles(self):
        return {
            "font": {
                "bold": Font(name="Arial", size=10, bold=True),
                "bold_red": Font(name="Arial", size=10,
                                 color=colors.RED, bold=True),
            },
            "fill": {
                "red": PatternFill("solid", fgColor="FF0000"),
                "grey": PatternFill("solid", fgColor="DDDDDD"),
                "yellow": PatternFill("solid", fgColor="FFFCB7"),
                "blue": PatternFill("solid", fgColor="9BF3FF"),
                "green": PatternFill("solid", fgColor="B0FF99"),
            },
            "align": {
                "left": Alignment(horizontal="left"),
                "center": Alignment(horizontal="center"),
                "right": Alignment(horizontal="right"),
            },
            "style": {
                "number": "#,##0.00",
                "date[dd/mm/yyyy]": "dd/mm/yyyy",
                "date[mm/dd/yyyy]": "mm/dd/yyyy",
                "datestamp": "yyyy-mm-dd",
                "text": "@",
                "percent": "0.00%",
            },
        }

    @api.model
    def create_xlsx_report(self, template, res_model, res_id):
        decoded_data = base64.decodestring(template.file)
        out_name = template.file_name
        obj_ir_config = self.env["ir.config_parameter"]
        ptemp = obj_ir_config.sudo().get_param("path_temp_file") or "/tmp"
        stamp = dt.utcnow().strftime("%H%M%S%f")[:-3]
        ftemp = "%s/temp%s.xlsx" % (ptemp, stamp)
        f = open(ftemp, "wb")
        f.write(decoded_data)
        f.seek(0)
        f.close()
        wb = load_workbook(ftemp)
        os.remove(ftemp)
        record = res_model and self.env[res_model].browse(res_id) or False
        self.generate_xlsx_report(wb, record, template)
        content = BytesIO()
        wb.save(content)
        content.seek(0)
        out_file = base64.encodestring(content.read())
        if record and 'name' in record and record.name:
            out_name = record.name.replace(" ", "").replace("/", "")
        else:
            fname = out_name.replace(" ", "").replace("/", "")
            ts = fields.Datetime.context_timestamp(self, dt.now())
            out_name = "%s_%s" % (fname, ts.strftime("%Y%m%d_%H%M%S"))
        if not out_name or len(out_name) == 0:
            out_name = "unknown"
        out_ext = "xlsx"
        return (out_file, "%s.%s" % (out_name, out_ext))

    @api.model
    def get_field_data(self, field_name, data):
        if not field_name:
            return None
        result = data
        for f in field_name.split("."):
            result = result[f]
        if isinstance(result, str):
            result = result.encode("utf-8")
        return result

    def set_style(self, field, field_style, styles):
        field_styles = field_style.split(";")
        for f in field_styles:
            (key, value) = f.split("=")
            if key not in styles.keys():
                raise UserError(_("Invalid style type %s" % key))
            if value.lower() not in styles[key].keys():
                raise UserError(
                    _("Invalid value %s for style type %s" % (value, key)))
            cell_style = styles[key][value]
            if key == "font":
                field.font = cell_style
            if key == "fill":
                field.fill = cell_style
            if key == "align":
                field.alignment = cell_style
            if key == "style":
                if value == "text":
                    try:
                        field.value = field.value.encode("utf-8")
                    except Exception:
                        field.value = str(field.value)
                field.number_format = cell_style

    @api.model
    def split_row_col(self, pos):
        match = re.match(r"([a-z]+)([0-9]+)", pos, re.I)
        if not match:
            raise UserError(_("Position %s is not valid") % pos)
        col, row = match.groups()
        return col, int(row)

    @api.model
    def generate_xlsx_report(self, workbook, data, template):
        for template_sheet in template.template_sheet_ids:
            sheet = workbook.get_sheet_by_name(template_sheet.name)
            for header in template_sheet.template_sheet_header_ids:
                header.set_header(sheet, data)
            for sheet_detail in template_sheet.template_sheet_detail_ids:
                line_field = self.get_field_data(
                    sheet_detail.field_name,
                    data,
                )
                if line_field:
                    i = 0
                    for line in line_field:
                        for detail in sheet_detail.detail_ids:
                            detail.set_detail(sheet, line, i)
                        i += 1
